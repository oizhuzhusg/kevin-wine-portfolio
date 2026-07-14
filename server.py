#!/usr/bin/env python3
import csv
import io
import json
import os
import re
import sqlite3
from datetime import date
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "data" / "kevin-wine.sqlite"
SCHEMA_PATH = ROOT / "schema.sql"
CONFIG_PATH = ROOT / "config" / "recommendation.json"
STATIC_DIR = ROOT / "static"

CATEGORIES = ["Drink Now", "Cellar", "Business", "Discovery"]
COLORS = ["red", "white", "sparkling", "sweet", "fortified"]

WATCHLIST = {
    "Burgundy Red": ["Arnaud Mortet", "Denis Mortet", "Jean Grivot", "Méo-Camuzet", "Hubert Lignier", "Rossignol-Trapet", "Joseph Drouhin", "Arlaud", "d’Angerville"],
    "Burgundy White": ["Hubert Lamy", "Pierre-Yves Colin-Morey", "Domaine Leflaive"],
    "Loire": ["Guiberteau", "Huet", "Thibaud Boudignon", "Château Yvonne", "Charles Joguet", "Philippe Alliet", "Belargus", "François Cotat", "Pascal Cotat"],
    "Rhône": ["Alain Graillot", "Gonon", "Clape", "Jean-Louis Chave", "Jamet"],
    "Bordeaux": ["Les Carmes Haut-Brion", "Château Canon", "Château Figeac", "Domaine de Chevalier", "Léoville Las Cases"],
    "Napa / USA": ["Chateau Montelena", "Dominus", "Corison", "Ridge Monte Bello"],
    "Italy": ["Sandrone", "Roagna", "Giacomo Conterno", "Bartolo Mascarello", "Giuseppe Mascarello"],
    "Spain": ["La Rioja Alta", "López de Heredia", "Vega Sicilia"]
}

PORTFOLIO_TARGETS = [
    ("Denis Mortet", "Gevrey-Chambertin Mes Cinq Terroirs", "Burgundy", "red", "2018-2020, 2022-2024", "2021", 140, 180, "Daily Drink", "Hold", "Tasted", 8.5, "yes", "红果、少量香草和橡木；酸度不突出，整体平衡。初开略有酒精感，瓶醒约30分钟后融合；单宁细腻。当前低于 Drouhin Cazetiers 2019。"),
    ("Arnaud Mortet", "Gevrey-Chambertin Ma Cuvée", "Burgundy", "red", "2018-2020, 2022-2024", "2021", 140, 180, "Daily Drink", "Hold", "Wishlist", None, None, "Gevrey 的果味与完成度坐标。"),
    ("Domaine Jean Grivot", "Vosne-Romanée", "Burgundy", "red", "2009, 2015-2020, 2022", "2021", 220, 280, "Discovery", "Ready", "Wishlist", None, None, "下一位重点验证的 Vosne 生产者。"),
    ("Hubert Lignier", "Morey-Saint-Denis Trilogie", "Burgundy", "red", "2015-2020, 2022", "2021", 130, 180, "Discovery", "Ready", "Wishlist", None, None, "用来认识 Morey-Saint-Denis 的平衡与结构。"),
    ("Hubert Lignier", "Morey-Saint-Denis 1er Cru Les Chaffots", "Burgundy", "red", "2016, 2019-2020", "2021", 150, 210, "Cellar", "Hold", "Wishlist", None, None, "Morey 的进阶对照。"),
    ("Rossignol-Trapet", "Gevrey-Chambertin", "Burgundy", "red", "2009, 2015-2017, 2019-2020, 2022", "2021", 110, 150, "Discovery", "Ready", "Wishlist", None, None, "与 Mortet 对照的传统 Gevrey。"),
    ("Domaine Fourrier", "Gevrey-Chambertin Vieille Vigne", "Burgundy", "red", "2014-2020", "2021", 250, 350, "Discovery", "Ready", "Wishlist", None, None, "测试纯净红果与细腻质地。"),
    ("Domaine Arlaud", "Morey-Saint-Denis", "Burgundy", "red", "2015-2020, 2022", "2021", 180, 220, "Discovery", "Ready", "Wishlist", None, None, "Morey 的果味清晰、质地细的坐标。"),
    ("Domaine Denis Bachelet", "Gevrey-Chambertin Vieilles Vignes", "Burgundy", "red", "2014-2017, 2019-2020", "2021", 270, 375, "Cellar", "Hold", "Wishlist", None, None, "浓度和精细度兼具的 Gevrey。"),
    ("Méo-Camuzet", "Vosne-Romanée", "Burgundy", "red", "2009, 2014-2020, 2022", "2021", 240, 320, "Discovery", "Hold", "Wishlist", None, None, "更丰润、有质感的 Vosne 对照。"),
    ("Hudelot-Noëllat", "Vosne-Romanée", "Burgundy", "red", "2014-2018, 2020", "2021", 195, 290, "Discovery", "Ready", "Wishlist", None, None, "与 Grivot 横向比较的 Vosne。"),
    ("Domaine Sylvie Esmonin", "Gevrey-Chambertin", "Burgundy", "red", "2009, 2014-2017, 2019", "2021", 140, 215, "Discovery", "Ready", "Wishlist", None, None, "更有力量与成熟感的 Gevrey。"),
    ("Ghislaine Barthod", "Chambolle-Musigny", "Burgundy", "red", "2015-2019, 2022", "2021", 180, 260, "Discovery", "Ready", "Tasted", 7.5, "maybe", "红果，花香不明显；偏酸，整体平衡但比较平淡，余味不长。当前排在 Denis Mortet Mes Cinq Terroirs 2023 之后；不考虑买整箱。"),
    ("Comte Armand", "Pommard 1er Cru Clos des Epeneaux", "Burgundy", "red", "2009, 2016, 2018-2020", "2021", 250, 340, "Heritage", "Ready", "Wishlist", None, None, "成熟年份优先，用来认识 Pommard。"),
    ("Marquis d'Angerville", "Volnay 1er Cru Taillepieds", "Burgundy", "red", "2009, 2016, 2018-2020", "2021", 260, 340, "Heritage", "Ready", "Wishlist", None, None, "成熟 Volnay 的标杆性对照。"),
    ("Hubert Lamy", "Saint-Aubin 1er Cru En Remilly", "Burgundy", "white", "2010, 2021-2023", "", 180, 250, "Discovery", "Ready", "Tasted", None, None, "2010 第一印象：酸度偏高。保留为生产者、地块与年份的独立记录。"),
    ("Hubert Lamy", "Saint-Aubin Village / Les Frionnes", "Burgundy", "white", "2020, 2022-2023", "", 160, 230, "Discovery", "Ready", "Wishlist", None, None, "用较有果味和中段的酒款重新验证 Lamy。"),
    ("Domaine Ramonet", "Chassagne-Montrachet", "Burgundy", "white", "2014-2015, 2017-2019", "", 210, 305, "Discovery", "Ready", "Wishlist", None, None, "白勃艮第的首要验证对象。"),
    ("Domaine Antoine Jobard", "Meursault", "Burgundy", "white", "2014-2015, 2017-2019", "", 195, 295, "Discovery", "Ready", "Wishlist", None, None, "中段更丰厚的 Meursault 对照。"),
    ("Pierre-Yves Colin-Morey", "Chassagne-Montrachet / Saint-Aubin", "Burgundy", "white", "2014-2015, 2017-2019", "", 175, 285, "Discovery", "Ready", "Wishlist", None, None, "张力与浓度兼具的高阶验证。"),
    ("Fontaine-Gagnard", "Chassagne-Montrachet", "Burgundy", "white", "2014-2015, 2017-2019", "", 120, 195, "Discovery", "Ready", "Wishlist", None, None, "果味与酒体较友好的 Chassagne 入口。"),
    ("Ballot-Millot", "Meursault", "Burgundy", "white", "2014-2015, 2017-2019", "", 130, 215, "Discovery", "Ready", "Wishlist", None, None, "质地完整的 Meursault。"),
    ("Domaine Paul Pillot", "Chassagne-Montrachet", "Burgundy", "white", "2015, 2017-2019", "", 165, 255, "Discovery", "Ready", "Wishlist", None, None, "纯净且有中段的 Chassagne。"),
    ("Morey-Coffinet", "Chassagne-Montrachet", "Burgundy", "white", "2015, 2017-2020", "", 120, 195, "Discovery", "Ready", "Wishlist", None, None, "果味、桶感与矿物感兼顾。"),
    ("Etienne Sauzet", "Puligny-Montrachet", "Burgundy", "white", "2014-2015, 2017-2019", "", 175, 265, "Discovery", "Ready", "Wishlist", None, None, "Puligny 的经典坐标。"),
    ("François Carillon", "Puligny-Montrachet", "Burgundy", "white", "2014-2015, 2017-2019", "", 140, 225, "Discovery", "Ready", "Wishlist", None, None, "较优雅但仍有果味和长度的 Puligny。"),
    ("Domaine Leflaive", "Bourgogne Blanc / Puligny-Montrachet", "Burgundy", "white", "2014-2015, 2017-2019", "", 150, 490, "Discovery", "Ready", "Wishlist", None, None, "先从大区级或单瓶开始建立 Puligny 坐标。"),
]

GLOBAL_TARGET_ROWS = """
Loire|Huet|Vouvray Le Mont Demi-Sec|white|2020-2023||85|130|Daily Drink|Ready
Loire|Huet|Vouvray Le Haut-Lieu Sec|white|2020-2024||70|110|Discovery|Ready
Loire|Belargus|Ronceray|white|2020-2023||90|150|Discovery|Hold
Loire|Nicolas Joly|Coulée de Serrant|white|2018-2022||145|210|Discovery|Ready
Loire|Pascal Cotat|Sancerre Les Monts Damnés|white|2020-2023||125|180|Discovery|Ready
Loire|Domaine Vacheron|Sancerre Blanc|white|2021-2023||90|140|Daily Drink|Ready
Loire|Domaine de la Pépière|Muscadet Clos des Briords|white|2020-2023||40|70|Daily Drink|Ready
Loire|Philippe Alliet|Chinon Vieilles Vignes|red|2019-2022||70|110|Daily Drink|Ready
Loire|Philippe Alliet|Chinon Coteau de Noiré|red|2018-2021||110|160|Discovery|Ready
Loire|Yannick Amirault|Bourgueil Le Grand Clos|red|2019-2022||80|120|Discovery|Ready
Loire|Clos Rougeard|Saumur-Champigny Le Clos|red|2015-2019||450|700|Heritage|Ready

Rhône|Alain Graillot|Crozes-Hermitage|red|2019-2022||80|130|Daily Drink|Ready
Rhône|Bernard Gripa|Saint-Joseph Rouge|red|2020-2022||100|150|Discovery|Ready
Rhône|Pierre Gonon|Saint-Joseph|red|2020-2022||180|280|Discovery|Hold
Rhône|Auguste Clape|Cornas|red|2018-2021||240|380|Cellar|Hold
Rhône|Jean-Louis Chave|Hermitage|red|2015-2020||450|700|Business|Hold
Rhône|Domaine Jamet|Côte-Rôtie|red|2016-2020||350|550|Cellar|Hold
Rhône|Domaine du Vieux Télégraphe|Châteauneuf-du-Pape La Crau|red|2015-2020||150|230|Cellar|Ready
Rhône|Clos des Papes|Châteauneuf-du-Pape|red|2016-2020||180|280|Cellar|Hold
Rhône|Domaine Rostaing|Côte-Rôtie Ampodium|red|2016-2020||170|260|Discovery|Ready
Rhône|Jean-Louis Chave|Hermitage Blanc|white|2016-2020||350|520|Business|Hold
Rhône|Domaine Georges Vernay|Condrieu Coteau de Vernon|white|2018-2022||150|230|Discovery|Ready
Rhône|Domaine François Villard|Condrieu De Poncins|white|2020-2023||95|150|Discovery|Ready

Bordeaux|Les Carmes Haut-Brion|Pessac-Léognan|red|2018-2020||190|290|Business|Hold
Bordeaux|Château Canon|Saint-Émilion|red|2018-2020||250|400|Business|Hold
Bordeaux|Château Figeac|Saint-Émilion|red|2018-2020||350|520|Business|Hold
Bordeaux|Château Léoville Las Cases|Saint-Julien|red|2016, 2018-2020||300|500|Cellar|Hold
Bordeaux|Château Montrose|Saint-Estèphe|red|2016, 2018-2020||210|340|Cellar|Hold
Bordeaux|Château Pichon Baron|Pauillac|red|2016, 2018-2020||200|320|Business|Hold
Bordeaux|Château Pichon Comtesse|Pauillac|red|2016, 2018-2020||220|350|Business|Hold
Bordeaux|Château Domaine de Chevalier|Pessac-Léognan Rouge|red|2016, 2018-2021||110|180|Daily Drink|Ready
Bordeaux|Château Laroque|Saint-Émilion|red|2019-2021||60|100|Daily Drink|Ready
Bordeaux|Château Haut-Brion|Pessac-Léognan Rouge|red|2015-2020||700|1100|Business|Hold
Bordeaux|Château Domaine de Chevalier|Pessac-Léognan Blanc|white|2017-2021||110|180|Discovery|Ready
Bordeaux|Château Smith Haut Lafitte|Pessac-Léognan Blanc|white|2018-2022||150|230|Business|Ready

Champagne|Egly-Ouriet|Grand Cru Brut Tradition|sparkling|NV||155|250|Discovery|Ready
Champagne|Agrapart|Terroirs Grand Cru|sparkling|NV||130|190|Discovery|Ready
Champagne|Agrapart|Minéral Grand Cru|sparkling|2016-2019||220|340|Cellar|Hold
Champagne|Pierre Péters|Cuvée de Réserve|sparkling|NV||90|150|Daily Drink|Ready
Champagne|Bérêche|Brut Réserve|sparkling|NV||100|160|Discovery|Ready
Champagne|Chartogne-Taillet|Sainte Anne|sparkling|NV||90|140|Discovery|Ready
Champagne|Ulysse Collin|Les Pierrières|sparkling|2017-2020||300|480|Cellar|Hold
Champagne|Jacques Selosse|Initial|sparkling|NV||650|1000|Business|Ready
Champagne|Krug|Grande Cuvée|sparkling|NV||350|500|Business|Ready
Champagne|Louis Roederer|Cristal|sparkling|2014-2016||450|700|Business|Hold

Germany|Keller|Von der Fels Riesling|white|2022-2024||70|110|Daily Drink|Ready
Germany|Keller|Kirchspiel GG|white|2019-2023||220|320|Cellar|Hold
Germany|Keller|Morstein GG|white|2018-2023||350|550|Cellar|Hold
Germany|Dönnhoff|Estate Riesling|white|2022-2024||45|75|Daily Drink|Ready
Germany|Dönnhoff|Hermannshöhle GG|white|2019-2023||180|260|Cellar|Hold
Germany|Wittmann|Kirchspiel GG|white|2019-2023||130|210|Cellar|Hold
Germany|Wittmann|Morstein GG|white|2019-2023||150|240|Cellar|Hold
Germany|Schäfer-Fröhlich|Felseneck GG|white|2019-2023||180|280|Cellar|Hold
Germany|Emrich-Schönleber|Halenberg GG|white|2019-2023||140|220|Cellar|Hold
Germany|J.J. Prüm|Graacher Himmelreich Spätlese|white|2018-2023||100|170|Discovery|Ready
Germany|Egon Müller|Scharzhofberger Kabinett|white|2018-2023||300|500|Business|Hold
Germany|Leitz|Berg Schlossberg GG|white|2020-2023||115|180|Discovery|Hold

Italy|Giacomo Conterno|Barbera d'Alba Francia|red|2018-2022||150|230|Daily Drink|Ready
Italy|Giacomo Conterno|Barolo Francia|red|2013-2019||550|850|Cellar|Hold
Italy|Roagna|Langhe Rosso|red|2018-2022||140|220|Daily Drink|Ready
Italy|Roagna|Barbaresco Pajé|red|2014-2019||350|550|Cellar|Hold
Italy|Bartolo Mascarello|Barolo|red|2013-2019||550|850|Cellar|Hold
Italy|Giuseppe Mascarello|Barolo Monprivato|red|2013-2019||450|700|Cellar|Hold
Italy|Luciano Sandrone|Barolo Le Vigne|red|2013-2019||250|380|Cellar|Hold
Italy|Vietti|Barolo Castiglione|red|2018-2021||120|180|Daily Drink|Hold
Italy|Quintarelli|Rosso del Bepi|red|2008-2014||350|550|Heritage|Ready
Italy|Dal Forno Romano|Valpolicella Superiore|red|2013-2018||280|450|Business|Hold
Italy|Biondi-Santi|Brunello di Montalcino|red|2012-2018||350|550|Cellar|Hold
Italy|Fontodi|Flaccianello della Pieve|red|2016-2020||170|260|Cellar|Hold
Italy|Isole e Olena|Cepparello|red|2016-2020||150|230|Cellar|Hold
Italy|Passopisciaro|Etna Rosso Contrada Guardiola|red|2017-2021||110|180|Discovery|Ready
Italy|Elena Walch|Beyond the Clouds|white|2018-2022||80|130|Discovery|Ready
Italy|Terlano|Terlaner I Primo Grande Cuvée|white|2016-2021||150|230|Cellar|Hold
Italy|Valentini|Trebbiano d'Abruzzo|white|2012-2018||350|550|Business|Hold
Italy|Emidio Pepe|Trebbiano d'Abruzzo|white|2010-2018||150|240|Heritage|Ready

Spain|Vega Sicilia|Único|red|2009-2014||500|800|Business|Ready
Spain|Vega Sicilia|Valbuena 5°|red|2014-2019||220|350|Business|Ready
Spain|La Rioja Alta|Gran Reserva 904|red|2010-2016||120|190|Heritage|Ready
Spain|La Rioja Alta|Gran Reserva 890|red|2004-2010||300|480|Heritage|Ready
Spain|López de Heredia|Viña Tondonia Reserva|red|2009-2012||80|130|Daily Drink|Ready
Spain|Marqués de Murrieta|Castillo Ygay Gran Reserva Especial|red|2010-2015||300|480|Business|Hold
Spain|Muga|Prado Enea Gran Reserva|red|2010-2016||100|160|Heritage|Ready
Spain|Álvaro Palacios|Finca Dofí|red|2015-2020||180|280|Cellar|Hold
Spain|Clos Mogador|Priorat|red|2010-2018||150|240|Heritage|Ready
Spain|Terroir al Límit|Arbossar|red|2015-2020||160|250|Discovery|Hold
Spain|Pingus|PSI|red|2018-2021||90|150|Discovery|Ready
Spain|Dominio de Pingus|Flor de Pingus|red|2016-2021||170|270|Cellar|Hold
Spain|Artadi|Viña El Pisón|red|2014-2019||280|440|Cellar|Hold
Spain|López de Heredia|Viña Tondonia Blanco Reserva|white|2008-2012||100|170|Discovery|Ready
Spain|Raúl Pérez|Ultreia La Claudina|white|2018-2022||90|150|Discovery|Ready

USA|Dominus Estate|Dominus|red|2013-2019||350|550|Business|Hold
USA|Ridge Vineyards|Monte Bello|red|2012-2019||300|500|Cellar|Hold
USA|Corison|Kronos Vineyard Cabernet Sauvignon|red|2013-2019||250|400|Cellar|Hold
USA|Chateau Montelena|Estate Cabernet Sauvignon|red|2013-2019||200|320|Cellar|Hold
USA|Heitz Cellar|Martha's Vineyard Cabernet Sauvignon|red|2012-2018||350|550|Cellar|Hold
USA|Dunn Vineyards|Howell Mountain Cabernet Sauvignon|red|2012-2018||250|400|Cellar|Hold
USA|Mayacamas|Cabernet Sauvignon|red|2012-2018||180|300|Cellar|Hold
USA|Spottswoode|Cabernet Sauvignon|red|2014-2019||250|380|Business|Hold
USA|Shafer|Hillside Select|red|2013-2018||350|550|Business|Hold
USA|Domaine Drouhin Oregon|Laurène Pinot Noir|red|2015-2020||120|190|Discovery|Ready
USA|Beaux Frères|The Beaux Frères Vineyard Pinot Noir|red|2015-2020||180|290|Discovery|Ready
USA|The Eyrie Vineyards|South Block Reserve Pinot Noir|red|2014-2020||150|240|Discovery|Ready
USA|Kistler|Les Noisetiers Chardonnay|white|2017-2022||120|190|Discovery|Ready
USA|Aubert|Larry Hyde & Sons Chardonnay|white|2017-2022||220|340|Business|Ready
USA|Mount Eden Vineyards|Estate Chardonnay|white|2015-2020||120|190|Discovery|Ready
USA|Ceritas|Charles Heintz Vineyard Chardonnay|white|2017-2022||150|240|Discovery|Ready

South America|Catena Zapata|Adrianna White Stones Chardonnay|white|2018-2022||180|280|Discovery|Ready
South America|Catena Zapata|Adrianna White Bones Chardonnay|white|2018-2022||200|320|Discovery|Ready
South America|Zuccardi|Finca Piedra Infinita|red|2017-2021||130|220|Discovery|Hold
South America|Zuccardi|Piedra Infinita Supercal|red|2017-2021||180|280|Cellar|Hold
South America|Per Se|Volare del Camino|red|2017-2021||180|300|Discovery|Hold
South America|El Enemigo|Gran Enemigo Gualtallary|red|2017-2021||100|170|Discovery|Ready
South America|Bodega Chacra|Cincuenta y Cinco Pinot Noir|red|2017-2022||110|180|Discovery|Ready
South America|Achával-Ferrer|Finca Altamira Malbec|red|2016-2020||120|190|Discovery|Ready
South America|Seña|Seña|red|2015-2020||190|300|Business|Hold
South America|Viñedo Chadwick|Cabernet Sauvignon|red|2015-2020||300|480|Business|Hold
South America|De Martino|Vigno Carignan|red|2017-2022||65|110|Daily Drink|Ready
South America|Clos Apalta|Clos Apalta|red|2015-2020||170|270|Cellar|Hold
South America|Errázuriz|Las Pizarras Chardonnay|white|2017-2022||100|170|Discovery|Ready
South America|Casa Marin|Cipreses Vineyard Sauvignon Blanc|white|2019-2023||75|120|Discovery|Ready
South America|Bodega Garzón|Balasto|red|2018-2022||100|170|Discovery|Ready

Australia|Cullen|Diana Madeline|red|2012, 2014-2016, 2018||155|250|Discovery|Ready
Australia|Henschke|Mount Edelstone Shiraz|red|2009, 2012-2016||195|310|Cellar|Hold
Australia|Mount Mary|Quintet|red|2010, 2012, 2015-2017||175|270|Discovery|Ready
Australia|Clonakilla|Shiraz Viognier|red|2012-2018||145|220|Discovery|Ready
Australia|Penfolds|St Henri Shiraz|red|2008, 2010, 2012, 2014-2016||125|200|Heritage|Ready
Australia|Yarra Yering|Dry Red No. 1|red|2012, 2014-2017||125|190|Discovery|Ready
Australia|Yangarra|High Sands Grenache|red|2015-2020||165|260|Discovery|Ready
Australia|Tolpuddle|Pinot Noir|red|2015-2019||125|190|Discovery|Ready
Australia|Bass Phillip|Premium Pinot Noir|red|2012-2017||195|320|Discovery|Ready
Australia|Torbreck|RunRig|red|2006, 2008-2013||280|445|Business|Hold
Australia|Leeuwin Estate|Art Series Chardonnay|white|2013, 2015-2019||145|220|Discovery|Ready
Australia|Giaconda|Estate Vineyard Chardonnay|white|2012, 2014-2017||240|380|Business|Hold
Australia|Cullen|Kevin John Chardonnay|white|2014-2018||175|270|Discovery|Ready
Australia|Tyrrell's|Vat 1 Semillon|white|2009, 2011, 2013-2014, 2017||105|170|Discovery|Ready
Australia|Pierro|Chardonnay|white|2013, 2015-2018||95|150|Discovery|Ready

New Zealand|Te Mata|Coleraine|red|2013-2016, 2018||125|200|Discovery|Ready
New Zealand|Ata Rangi|Pinot Noir|red|2013, 2015-2018||115|180|Discovery|Ready
New Zealand|Felton Road|Block 5 Pinot Noir|red|2014-2018||155|250|Discovery|Ready
New Zealand|Kusuda|Pinot Noir|red|2013-2017||205|320|Discovery|Ready
New Zealand|Burn Cottage|Pinot Noir|red|2014-2018||105|160|Discovery|Ready
New Zealand|Escarpment|Kupe Pinot Noir|red|2013-2017||105|170|Discovery|Ready
New Zealand|Craggy Range|Le Sol Syrah|red|2013-2017||145|220|Discovery|Ready
New Zealand|Te Mata|Bullnose Syrah|red|2014-2018||85|130|Daily Drink|Ready
New Zealand|Trinity Hill|Homage Syrah|red|2013-2017||125|200|Discovery|Ready
New Zealand|Prophet's Rock|Pinot Noir|red|2015-2019||95|150|Discovery|Ready
New Zealand|Kumeu River|Mate's Vineyard Chardonnay|white|2014-2018||125|200|Discovery|Ready
New Zealand|Kumeu River|Hunting Hill Chardonnay|white|2015-2019||85|140|Discovery|Ready
New Zealand|Te Mata|Elston Chardonnay|white|2014-2018||65|110|Daily Drink|Ready
New Zealand|Neudorf|Moutere Chardonnay|white|2014-2018||95|150|Discovery|Ready
New Zealand|Dog Point|Section 94 Sauvignon Blanc|white|2014-2018||65|110|Discovery|Ready

China|Ao Yun|Ao Yun|red|2015-2020||380|660|Business|Hold
China|Long Dai|Long Dai|red|2017-2021||275|460|Business|Hold
China|Silver Heights|The Summit 阙歌|red|2013-2021||80|145|Discovery|Ready
China|Helan Qingxue|Jia Bei Lan Reserve 加贝兰珍藏|red|2012-2021||80|165|Discovery|Ready
China|Grace Vineyard|Chairman's Reserve 庄主珍藏|red|2012-2020||80|145|Discovery|Ready
China|Puchang Vineyard|Reserve Saperavi|red|2016-2022||60|115|Discovery|Ready
China|Canaan Winery|Black Beauty 黑美人|red|2016-2021||80|145|Discovery|Ready
China|Tiansai|Skyline of Gobi Selection|red|2016-2022||60|105|Daily Drink|Ready
China|Xige Estate|N.28 Cabernet Gernischt|red|2017-2022||50|95|Daily Drink|Ready
China|Domaine Franco-Chinois|Marselan Reserve 珍藏马瑟兰|red|2017-2022||60|115|Discovery|Ready
China|Silver Heights|Family Reserve Chardonnay|white|2018-2023||50|95|Discovery|Ready
China|Grace Vineyard|Symphony Chardonnay 奏鸣曲霞多丽|white|2018-2023||45|85|Daily Drink|Ready
China|Puchang Vineyard|Rkatsiteli 白羽|white|2018-2023||45|85|Discovery|Ready
China|Longting Vineyard|Petit Manseng 小芒森|white|2019-2023||60|105|Discovery|Ready
China|Canaan Winery|Mastery Chardonnay 诗百篇霞多丽|white|2018-2023||60|105|Discovery|Ready
"""

REGION_COUNTRIES = {
    "Loire": "France", "Rhône": "France", "Bordeaux": "France", "Champagne": "France",
    "Germany": "Germany", "Italy": "Italy", "Spain": "Spain", "USA": "United States",
    "South America": "Argentina / Chile / Uruguay", "Australia": "Australia", "New Zealand": "New Zealand", "China": "China",
}

RETIRED_TARGETS = [
    ("Domaine Leflaive", "Bourgogne Blanc / Puligny-Montrachet"),
    ("François Carillon", "Puligny-Montrachet"),
]


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def load_config():
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def normalize(s):
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def contains_name(text, name):
    return normalize(name) in normalize(text)


def now_year():
    return date.today().year


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    first = not DB_PATH.exists()
    with db() as conn:
        conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        if first or conn.execute("SELECT COUNT(*) FROM producer_watchlist").fetchone()[0] == 0:
            seed(conn)
        seed_portfolio_targets(conn)


def seed_portfolio_targets(conn):
    for producer, wine_name in RETIRED_TARGETS:
        conn.execute("DELETE FROM portfolio_targets WHERE producer = ? AND wine_name = ? AND status = 'Wishlist'", (producer, wine_name))
    for target in PORTFOLIO_TARGETS:
        conn.execute(
            """
            INSERT OR IGNORE INTO portfolio_targets
              (producer, wine_name, region, color, recommended_vintages, avoid_vintages,
               ideal_price_sgd, max_price_sgd, role, stage, status, personal_score,
               would_buy_again, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            target,
        )
    for raw_row in GLOBAL_TARGET_ROWS.strip().splitlines():
        if not raw_row.strip():
            continue
        region, producer, wine_name, color, vintages, avoid, ideal, maximum, role, stage = raw_row.split("|")
        conn.execute(
            """
            INSERT OR IGNORE INTO portfolio_targets
              (producer, wine_name, region, country, color, recommended_vintages, avoid_vintages,
               ideal_price_sgd, max_price_sgd, role, stage, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (producer, wine_name, region, REGION_COUNTRIES[region], color, vintages, avoid, int(ideal), int(maximum), role, stage, "按现阶段探索规则列入；新生产者先买一瓶验证。"),
        )
    for producer, wine_name in RETIRED_TARGETS:
        conn.execute("DELETE FROM portfolio_targets WHERE producer = ? AND wine_name = ? AND status = 'Wishlist'", (producer, wine_name))
    conn.commit()


def seed(conn):
    for group, producers in WATCHLIST.items():
        for producer in producers:
            color = "white" if "White" in group else "red" if group in ["Burgundy Red", "Rhône", "Bordeaux", "Napa / USA", "Italy", "Spain"] else None
            conn.execute(
                "INSERT OR IGNORE INTO producer_watchlist (producer, region_group, color_focus) VALUES (?, ?, ?)",
                (producer, group, color),
            )
    for category, pct in load_config()["category_targets"].items():
        conn.execute(
            "INSERT OR IGNORE INTO target_portfolio (category, target_percentage, notes) VALUES (?, ?, ?)",
            (category, pct, "Kevin Wine Portfolio default target"),
        )
    samples = [
        {
            "producer": "Arnaud Mortet", "wine_name": "Gevrey-Chambertin Mes Cinq Terroirs", "region": "Burgundy", "country": "France",
            "appellation": "Gevrey-Chambertin", "classification": "Village", "grape_variety": "Pinot Noir", "color": "red", "vintage": 2020,
            "drinking_window_start": 2025, "drinking_window_end": 2035, "ideal_price_sgd": 190, "max_price_sgd": 260,
            "personal_priority": "A", "target_inventory": 3, "current_inventory": 1, "categories": ["Cellar", "Business"], "styles": ["fruit_forward", "balanced", "structured"]
        },
        {
            "producer": "La Rioja Alta", "wine_name": "Gran Reserva 904", "region": "Rioja", "country": "Spain",
            "classification": "Gran Reserva", "grape_variety": "Tempranillo blend", "color": "red", "vintage": 2015,
            "drinking_window_start": 2023, "drinking_window_end": 2035, "ideal_price_sgd": 120, "max_price_sgd": 160,
            "personal_priority": "A", "target_inventory": 4, "current_inventory": 2, "categories": ["Drink Now", "Business"], "styles": ["balanced", "fruit_forward"]
        },
        {
            "producer": "Huet", "wine_name": "Vouvray Le Mont Sec", "region": "Loire", "country": "France",
            "appellation": "Vouvray", "grape_variety": "Chenin Blanc", "color": "white", "vintage": 2020,
            "drinking_window_start": 2024, "drinking_window_end": 2032, "ideal_price_sgd": 95, "max_price_sgd": 130,
            "personal_priority": "B", "target_inventory": 2, "current_inventory": 1, "categories": ["Discovery", "Cellar"], "styles": ["balanced", "mineral"]
        },
    ]
    for item in samples:
        categories = item.pop("categories")
        styles = item.pop("styles")
        columns = ",".join(item.keys())
        placeholders = ",".join(["?"] * len(item))
        cur = conn.execute(f"INSERT INTO wines ({columns}) VALUES ({placeholders})", tuple(item.values()))
        wine_id = cur.lastrowid
        for category in categories:
            conn.execute("INSERT INTO wine_category_tags (wine_id, category) VALUES (?, ?)", (wine_id, category))
        for style in styles:
            conn.execute("INSERT INTO wine_style_tags (wine_id, style_tag) VALUES (?, ?)", (wine_id, style))
        conn.execute(
            "INSERT INTO buy_list (wine_id, producer, wine_name, recommended_vintages, ideal_price_sgd, max_price_sgd, current_inventory, target_inventory, recommendation_grade, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (wine_id, item["producer"], item["wine_name"], "2016, 2018, 2019, 2020", item["ideal_price_sgd"], item["max_price_sgd"], item["current_inventory"], item["target_inventory"], item["personal_priority"], "Seed target"),
        )
    conn.commit()


def row_to_dict(row):
    return dict(row) if row else None


def wine_payload(conn, row):
    item = row_to_dict(row)
    if not item:
        return None
    item["category_tags"] = [r["category"] for r in conn.execute("SELECT category FROM wine_category_tags WHERE wine_id = ? ORDER BY category", (item["id"],))]
    item["style_tags"] = [r["style_tag"] for r in conn.execute("SELECT style_tag FROM wine_style_tags WHERE wine_id = ? ORDER BY style_tag", (item["id"],))]
    return item


def all_wines(conn):
    return [wine_payload(conn, r) for r in conn.execute("SELECT * FROM wines ORDER BY producer, wine_name, vintage DESC")]


def dashboard(conn):
    wines = all_wines(conn)
    total_bottles = sum(int(w["current_inventory"] or 0) for w in wines)
    color_counts = {c: 0 for c in COLORS}
    for w in wines:
        color_counts[w["color"]] = color_counts.get(w["color"], 0) + int(w["current_inventory"] or 0)
    cat_counts = {c: 0 for c in CATEGORIES}
    for w in wines:
        for c in w["category_tags"]:
            cat_counts[c] += int(w["current_inventory"] or 0)
    purchases = conn.execute("SELECT SUM(total_cost) AS total, SUM(quantity) AS qty FROM purchases").fetchone()
    total_cost = float(purchases["total"] or 0)
    avg_cost = total_cost / purchases["qty"] if purchases["qty"] else 0
    config = load_config()
    replenish = []
    for category, target in config["category_targets"].items():
        current = (cat_counts.get(category, 0) / total_bottles) if total_bottles else 0
        if current + 0.03 < target:
            replenish.append({"category": category, "current": current, "target": target})
    entering = [
        w for w in wines
        if w["current_inventory"] and w["drinking_window_start"] and now_year() <= int(w["drinking_window_start"]) <= now_year() + 2
    ]
    return {
        "total_bottles": total_bottles,
        "color_counts": color_counts,
        "color_percentages": {k: (v / total_bottles if total_bottles else 0) for k, v in color_counts.items()},
        "category_counts": cat_counts,
        "category_percentages": {k: (v / total_bottles if total_bottles else 0) for k, v in cat_counts.items()},
        "total_cost": total_cost,
        "average_bottle_cost": avg_cost,
        "replenish": replenish,
        "entering_window": entering[:8],
        "targets": config
    }


def parse_money(value):
    if value is None:
        return None
    text = str(value).replace(",", "")
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return float(match.group(1)) if match else None


def parse_int(value):
    if value in (None, ""):
        return None
    match = re.search(r"(19|20)\d{2}", str(value))
    return int(match.group(0)) if match else None


def header_value(row, names):
    normalized = {normalize(k): v for k, v in row.items()}
    for name in names:
        if normalize(name) in normalized:
            return normalized[normalize(name)]
    return None


def parse_price_rows(filename, content):
    rows = []
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("XLSX parsing requires openpyxl. Use the bundled Python runtime or upload CSV.") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(v or "").strip() for v in values[0]]
        for values_row in values[1:]:
            rows.append({headers[i]: values_row[i] if i < len(values_row) else None for i in range(len(headers))})
    else:
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.DictReader(io.StringIO(text)))
    parsed = []
    for row in rows:
        producer = header_value(row, ["producer", "domaine", "estate", "chateau", "château"])
        wine_name = header_value(row, ["wine_name", "wine", "name", "cuvee", "cuvée", "description"])
        raw_text = " ".join(str(v or "") for v in row.values())
        if not producer:
            producer = infer_producer(raw_text)
        if not wine_name:
            wine_name = raw_text
        parsed.append({
            "producer": str(producer or "").strip(),
            "wine_name": str(wine_name or "").strip(),
            "vintage": parse_int(header_value(row, ["vintage", "year"]) or raw_text),
            "region": str(header_value(row, ["region", "appellation", "country"]) or "").strip(),
            "color": normalize(header_value(row, ["color", "type"]) or ""),
            "price_sgd": parse_money(header_value(row, ["price_sgd", "price", "offer price", "sgd", "unit price"])),
            "raw_text": raw_text.strip()
        })
    return [r for r in parsed if r["wine_name"] or r["producer"]]


def infer_producer(text):
    for producers in WATCHLIST.values():
        for producer in producers:
            if contains_name(text, producer):
                return producer
    return ""


def known_wine_for_offer(conn, offer):
    haystack = f"{offer.get('producer','')} {offer.get('wine_name','')} {offer.get('raw_text','')}"
    for wine in all_wines(conn):
        if contains_name(haystack, wine["producer"]) and normalize(wine["wine_name"]).split(" ")[0:2]:
            name_bits = normalize(wine["wine_name"]).split()
            matched_bits = sum(1 for bit in name_bits[:4] if bit in normalize(haystack))
            if matched_bits >= min(2, len(name_bits)):
                return wine
    return None


def buy_list_for_offer(conn, offer):
    haystack = f"{offer.get('producer','')} {offer.get('wine_name','')} {offer.get('raw_text','')}"
    for row in conn.execute("SELECT * FROM buy_list"):
        item = row_to_dict(row)
        if contains_name(haystack, item["producer"]):
            bits = normalize(item["wine_name"]).split()
            if not bits or sum(1 for bit in bits[:4] if bit in normalize(haystack)) >= min(2, len(bits)):
                return item
    return None


def is_watchlist(conn, text):
    for row in conn.execute("SELECT producer FROM producer_watchlist"):
        if contains_name(text, row["producer"]):
            return row["producer"]
    return None


def recommend(conn, offer):
    cfg = load_config()
    score = 0
    reasons = []
    text = f"{offer.get('producer','')} {offer.get('wine_name','')} {offer.get('region','')} {offer.get('raw_text','')}"
    price = offer.get("price_sgd")
    known = known_wine_for_offer(conn, offer)
    target = buy_list_for_offer(conn, offer)
    watch = is_watchlist(conn, text)
    dash = dashboard(conn)
    ideal = None
    max_price = None
    categories = []
    current_inventory = 0
    target_inventory = 1
    color = offer.get("color") if offer.get("color") in COLORS else None

    if known:
        score += cfg["weights"]["known_wine_match"]
        ideal = known["ideal_price_sgd"]
        max_price = known["max_price_sgd"]
        categories = known["category_tags"]
        current_inventory = int(known["current_inventory"] or 0)
        target_inventory = int(known["target_inventory"] or 1)
        color = color or known["color"]
        reasons.append("已在库存/目标酒款中找到匹配")
    if target:
        score += cfg["weights"]["buy_list_match"]
        ideal = ideal or target["ideal_price_sgd"]
        max_price = max_price or target["max_price_sgd"]
        target_inventory = int(target["target_inventory"] or target_inventory)
        current_inventory = int(target["current_inventory"] or current_inventory)
        reasons.append("命中 Buy List")
    if watch:
        score += cfg["weights"]["watchlist"]
        reasons.append(f"核心生产者 watchlist：{watch}")
    for region in cfg["preferred_regions"]:
        if contains_name(text, region):
            score += cfg["weights"]["preferred_region"]
            reasons.append(f"符合偏好产区：{region}")
            break
    for tag in cfg["positive_style_tags"]:
        if contains_name(text, tag.replace("_", " ")):
            score += cfg["weights"]["positive_style"]
            reasons.append(f"风格正向：{tag}")
    for tag in cfg["negative_style_tags"]:
        if contains_name(text, tag.replace("_", " ")):
            score += cfg["weights"]["negative_style"]
            reasons.append(f"风格风险：{tag}")
    if price and ideal and price <= ideal:
        score += cfg["weights"]["below_ideal_price"]
        reasons.append("报价低于 ideal price")
    elif price and max_price and price <= max_price:
        score += cfg["weights"]["below_max_price"]
        reasons.append("报价低于 max price")
    elif price and max_price and price > max_price:
        score += cfg["weights"]["above_max_price"]
        reasons.append("报价高于 max price")
    if current_inventory < target_inventory:
        score += cfg["weights"]["inventory_gap"]
        reasons.append(f"库存低于目标：{current_inventory}/{target_inventory}")
    if color in cfg["color_targets"]:
        current_pct = dash["color_percentages"].get(color, 0)
        if current_pct < cfg["color_targets"][color]:
            score += cfg["weights"]["color_need"]
            reasons.append(f"{color} 当前比例低于目标")
    for category in categories:
        if category in cfg["category_targets"] and dash["category_percentages"].get(category, 0) < cfg["category_targets"][category]:
            score += cfg["weights"]["category_need"]
            reasons.append(f"{category} 分类有缺口")
            break
    if contains_name(text, "grand cru") and not watch:
        score += cfg["weights"]["weak_grand_cru_penalty"]
        reasons.append("Grand Cru 标签但生产者不在核心名单")
    for producer in cfg["financialized_producers"]:
        if contains_name(text, producer):
            score += cfg["weights"]["financialized_penalty"]
            reasons.append("金融化酒庄，默认不追高")
            break
    if price and price > cfg["main_budget_sgd"]["max"] and not any(c in cfg["premium_budget_requires"] for c in categories):
        score += cfg["weights"]["premium_without_reason_penalty"]
        reasons.append("超过主力预算且缺少 Business/Cellar 理由")
    thresholds = cfg["grade_thresholds"]
    grade = "Avoid"
    if score >= thresholds["S"]:
        grade = "S"
    elif score >= thresholds["A"]:
        grade = "A"
    elif score >= thresholds["B"]:
        grade = "B"
    elif score >= thresholds["C"]:
        grade = "C"
    suggested = 0
    if grade == "S":
        suggested = min(3, max(1, target_inventory - current_inventory))
    elif grade == "A":
        suggested = min(2, max(1, target_inventory - current_inventory))
    elif grade == "B":
        suggested = 1 if current_inventory < target_inventory else 0
    return {
        "producer": offer.get("producer"),
        "wine_name": offer.get("wine_name"),
        "vintage": offer.get("vintage"),
        "region": offer.get("region"),
        "color": color,
        "price_sgd": price,
        "target_price_sgd": ideal,
        "max_price_sgd": max_price,
        "recommendation_grade": grade,
        "suggested_quantity": suggested,
        "reasons": "；".join(reasons) if reasons else "信息不足，先列为观察",
        "score": score,
        "raw_text": offer.get("raw_text", "")
    }


class Handler(BaseHTTPRequestHandler):
    server_version = "KevinWinePortfolio/0.1"

    def do_GET(self):
        try:
            self.route_get()
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self):
        try:
            self.route_post()
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_PATCH(self):
        try:
            self.route_patch()
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def route_get(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/"):
            with db() as conn:
                if path == "/api/dashboard":
                    return self.send_json(dashboard(conn))
                if path == "/api/wines":
                    return self.send_json(all_wines(conn))
                if path == "/api/purchases":
                    rows = [
                        row_to_dict(r)
                        for r in conn.execute(
                            """
                            SELECT purchases.*, wines.producer, wines.wine_name, wines.vintage
                            FROM purchases
                            JOIN wines ON wines.id = purchases.wine_id
                            ORDER BY purchases.purchase_date DESC, purchases.id DESC
                            """
                        )
                    ]
                    return self.send_json(rows)
                if path == "/api/buy-list":
                    rows = [row_to_dict(r) for r in conn.execute("SELECT * FROM buy_list ORDER BY recommendation_grade, producer")]
                    return self.send_json(rows)
                if path == "/api/portfolio-targets":
                    rows = [row_to_dict(r) for r in conn.execute("SELECT * FROM portfolio_targets ORDER BY color, producer, wine_name")]
                    return self.send_json(rows)
                if path == "/api/lookups":
                    return self.send_json({
                        "categories": CATEGORIES,
                        "colors": COLORS,
                        "watchlist": [row_to_dict(r) for r in conn.execute("SELECT * FROM producer_watchlist ORDER BY region_group, producer")]
                    })
                match = re.match(r"^/api/analysis/(\d+)/export\.csv$", path)
                if match:
                    return self.export_analysis(conn, int(match.group(1)))
            return self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
        self.serve_static(path)

    def route_post(self):
        parsed = urlparse(self.path)
        path = parsed.path
        with db() as conn:
            if path == "/api/wines":
                payload = self.read_json()
                wine_id = create_wine(conn, payload)
                return self.send_json(wine_payload(conn, conn.execute("SELECT * FROM wines WHERE id = ?", (wine_id,)).fetchone()), HTTPStatus.CREATED)
            if path == "/api/purchases":
                payload = self.read_json()
                total = float(payload.get("total_cost") or (float(payload.get("price_sgd", 0)) * int(payload.get("quantity", 1)) + float(payload.get("delivery_fee") or 0)))
                conn.execute(
                    "INSERT INTO purchases (wine_id, purchase_date, merchant, price_sgd, quantity, tax_included, delivery_fee, total_cost, purchase_reason, source_file_or_link) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (payload["wine_id"], payload["purchase_date"], payload.get("merchant"), payload["price_sgd"], payload["quantity"], int(bool(payload.get("tax_included", True))), payload.get("delivery_fee") or 0, total, payload.get("purchase_reason"), payload.get("source_file_or_link")),
                )
                conn.execute("UPDATE wines SET current_inventory = current_inventory + ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (int(payload["quantity"]), int(payload["wine_id"])))
                conn.commit()
                return self.send_json({"ok": True})
            if path == "/api/tastings":
                payload = self.read_json()
                conn.execute(
                    "INSERT INTO tastings (wine_id, tasting_date, occasion, decanting_time, serving_temperature, aroma_notes, palate_notes, structure_notes, food_pairing, personal_score, would_buy_again, preferred_use_case_after_tasting, tasting_summary) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (payload["wine_id"], payload["tasting_date"], payload.get("occasion"), payload.get("decanting_time"), payload.get("serving_temperature"), payload.get("aroma_notes"), payload.get("palate_notes"), payload.get("structure_notes"), payload.get("food_pairing"), payload.get("personal_score"), payload.get("would_buy_again"), payload.get("preferred_use_case_after_tasting"), payload.get("tasting_summary")),
                )
                conn.execute("UPDATE wines SET tasted_before = 1, personal_score = COALESCE(?, personal_score), current_inventory = MAX(current_inventory - ?, 0), updated_at = CURRENT_TIMESTAMP WHERE id = ?", (payload.get("personal_score"), int(payload.get("consume_quantity", 1)), int(payload["wine_id"])))
                conn.commit()
                return self.send_json({"ok": True})
            if path == "/api/buy-list":
                payload = self.read_json()
                conn.execute(
                    "INSERT INTO buy_list (wine_id, producer, wine_name, recommended_vintages, ideal_price_sgd, max_price_sgd, current_inventory, target_inventory, recommendation_grade, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (payload.get("wine_id"), payload["producer"], payload["wine_name"], payload.get("recommended_vintages"), payload.get("ideal_price_sgd"), payload.get("max_price_sgd"), payload.get("current_inventory") or 0, payload.get("target_inventory") or 1, payload.get("recommendation_grade") or "B", payload.get("notes")),
                )
                conn.commit()
                return self.send_json({"ok": True}, HTTPStatus.CREATED)
            if path == "/api/portfolio-targets":
                payload = self.read_json()
                fields = ["producer", "wine_name", "region", "country", "color", "recommended_vintages", "avoid_vintages", "ideal_price_sgd", "max_price_sgd", "role", "stage", "status", "personal_score", "would_buy_again", "notes"]
                data = {field: payload.get(field) for field in fields}
                data["producer"] = data["producer"] or "Unknown"
                data["wine_name"] = data["wine_name"] or "Unnamed wine"
                data["color"] = data["color"] or "red"
                data["role"] = data["role"] or "Discovery"
                data["stage"] = data["stage"] or "Ready"
                data["status"] = data["status"] or "Wishlist"
                columns = ",".join(data.keys())
                placeholders = ",".join(["?"] * len(data))
                target_id = conn.execute(f"INSERT INTO portfolio_targets ({columns}) VALUES ({placeholders})", tuple(data.values())).lastrowid
                conn.commit()
                return self.send_json(row_to_dict(conn.execute("SELECT * FROM portfolio_targets WHERE id = ?", (target_id,)).fetchone()), HTTPStatus.CREATED)
            if path == "/api/analyze-price-list":
                filename, content = self.read_upload()
                offers = parse_price_rows(filename, content)
                run = conn.execute("INSERT INTO analysis_runs (source_name) VALUES (?)", (filename,)).lastrowid
                results = []
                for offer in offers:
                    rec = recommend(conn, offer)
                    results.append(rec)
                    conn.execute(
                        "INSERT INTO analysis_items (run_id, producer, wine_name, vintage, region, color, price_sgd, target_price_sgd, max_price_sgd, recommendation_grade, suggested_quantity, reasons, raw_text) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (run, rec["producer"], rec["wine_name"], rec["vintage"], rec["region"], rec["color"], rec["price_sgd"], rec["target_price_sgd"], rec["max_price_sgd"], rec["recommendation_grade"], rec["suggested_quantity"], rec["reasons"], rec["raw_text"]),
                    )
                conn.commit()
                return self.send_json({"run_id": run, "items": results})
        return self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)

    def route_patch(self):
        path = urlparse(self.path).path
        target_match = re.match(r"^/api/portfolio-targets/(\d+)$", path)
        if target_match:
            payload = self.read_json()
            allowed = {"producer", "wine_name", "region", "country", "color", "recommended_vintages", "avoid_vintages", "ideal_price_sgd", "max_price_sgd", "role", "stage", "status", "personal_score", "would_buy_again", "notes"}
            updates = {key: value for key, value in payload.items() if key in allowed}
            if not updates:
                return self.send_json({"error": "No editable fields supplied"}, HTTPStatus.BAD_REQUEST)
            updates["updated_at"] = "CURRENT_TIMESTAMP"
            assignments = ", ".join(f"{field} = ?" for field in updates if field != "updated_at") + ", updated_at = CURRENT_TIMESTAMP"
            values = [value for field, value in updates.items() if field != "updated_at"] + [int(target_match.group(1))]
            with db() as conn:
                conn.execute(f"UPDATE portfolio_targets SET {assignments} WHERE id = ?", values)
                conn.commit()
                row = conn.execute("SELECT * FROM portfolio_targets WHERE id = ?", (int(target_match.group(1)),)).fetchone()
                if not row:
                    return self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
                return self.send_json(row_to_dict(row))
        match = re.match(r"^/api/wines/(\d+)$", path)
        if not match:
            return self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
        payload = self.read_json()
        wine_id = int(match.group(1))
        with db() as conn:
            update_wine(conn, wine_id, payload)
            return self.send_json(wine_payload(conn, conn.execute("SELECT * FROM wines WHERE id = ?", (wine_id,)).fetchone()))

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        return json.loads(self.rfile.read(length).decode("utf-8") or "{}")

    def read_upload(self):
        content_type = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        if "multipart/form-data" not in content_type:
            payload = json.loads(body.decode("utf-8"))
            return payload.get("filename", "pasted.csv"), payload.get("text", "").encode("utf-8")
        boundary = content_type.split("boundary=", 1)[1].encode()
        parts = body.split(b"--" + boundary)
        for part in parts:
            if b'name="file"' in part:
                header, data = part.split(b"\r\n\r\n", 1)
                data = data.rsplit(b"\r\n", 1)[0]
                name_match = re.search(br'filename="([^"]+)"', header)
                filename = name_match.group(1).decode("utf-8", errors="replace") if name_match else "upload.csv"
                return filename, data
        raise ValueError("No file field found")

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def export_analysis(self, conn, run_id):
        rows = [row_to_dict(r) for r in conn.execute("SELECT producer, wine_name, vintage, price_sgd, target_price_sgd, max_price_sgd, recommendation_grade, suggested_quantity, reasons FROM analysis_items WHERE run_id = ? ORDER BY CASE recommendation_grade WHEN 'S' THEN 1 WHEN 'A' THEN 2 WHEN 'B' THEN 3 WHEN 'C' THEN 4 ELSE 5 END, price_sgd", (run_id,))]
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=["producer", "wine_name", "vintage", "price_sgd", "target_price_sgd", "max_price_sgd", "recommendation_grade", "suggested_quantity", "reasons"])
        writer.writeheader()
        writer.writerows(rows)
        data = out.getvalue().encode("utf-8-sig")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", f'attachment; filename="kevin-wine-analysis-{run_id}.csv"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_static(self, path):
        if path == "/":
            path = "/index.html"
        target = (STATIC_DIR / unquote(path.lstrip("/"))).resolve()
        if not str(target).startswith(str(STATIC_DIR)) or not target.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        content_type = "text/html; charset=utf-8"
        if target.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif target.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        data = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def create_wine(conn, payload):
    fields = ["producer", "wine_name", "region", "country", "appellation", "vineyard_or_climat", "classification", "grape_variety", "color", "vintage", "bottle_size", "alcohol", "drinking_window_start", "drinking_window_end", "ideal_price_sgd", "max_price_sgd", "current_market_price_sgd", "personal_priority", "target_inventory", "current_inventory", "tasted_before", "personal_score", "notes"]
    data = {f: payload.get(f) for f in fields}
    data["color"] = data["color"] or "red"
    data["producer"] = data["producer"] or "Unknown"
    data["wine_name"] = data["wine_name"] or "Unnamed wine"
    columns = ",".join(data.keys())
    placeholders = ",".join(["?"] * len(data))
    wine_id = conn.execute(f"INSERT INTO wines ({columns}) VALUES ({placeholders})", tuple(data.values())).lastrowid
    replace_tags(conn, wine_id, payload.get("category_tags") or ["Discovery"], payload.get("style_tags") or [])
    conn.commit()
    return wine_id


def update_wine(conn, wine_id, payload):
    tag_fields = {"category_tags", "style_tags"}
    fields = [k for k in payload.keys() if k not in tag_fields]
    if fields:
        sets = ", ".join([f"{f} = ?" for f in fields])
        conn.execute(f"UPDATE wines SET {sets}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", tuple(payload[f] for f in fields) + (wine_id,))
    if "category_tags" in payload or "style_tags" in payload:
        current = wine_payload(conn, conn.execute("SELECT * FROM wines WHERE id = ?", (wine_id,)).fetchone())
        replace_tags(conn, wine_id, payload.get("category_tags", current["category_tags"]), payload.get("style_tags", current["style_tags"]))
    conn.commit()


def replace_tags(conn, wine_id, categories, styles):
    conn.execute("DELETE FROM wine_category_tags WHERE wine_id = ?", (wine_id,))
    conn.execute("DELETE FROM wine_style_tags WHERE wine_id = ?", (wine_id,))
    for category in categories:
        if category in CATEGORIES:
            conn.execute("INSERT OR IGNORE INTO wine_category_tags (wine_id, category) VALUES (?, ?)", (wine_id, category))
    for style in styles:
        style = normalize(style).replace(" ", "_")
        if style:
            conn.execute("INSERT OR IGNORE INTO wine_style_tags (wine_id, style_tag) VALUES (?, ?)", (wine_id, style))


def main():
    init_db()
    port = int(os.environ.get("PORT", "5188"))
    httpd = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Kevin Wine Portfolio running at http://localhost:{port}")
    httpd.serve_forever()


if __name__ == "__main__":
    main()
